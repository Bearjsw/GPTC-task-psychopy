# -*- coding: utf-8 -*-
"""0818_product_list.xlsx 를 stim_0818/ 아래 CSV로 옮긴다.

엑셀이 원본이다. 문구를 고쳤으면 엑셀을 고치고 이걸 다시 돌린다.

    python tools_import_0818.py                       기본 경로에서 찾는다
    python tools_import_0818.py 경로/파일.xlsx        직접 지정

엑셀 열 구성
    product_EN  product_KR  category  low  high
    UT_1 UT_2 UT_3  PhaseII_UT  HE_1 HE_2 HE_3  PhaseII_HE

내보내는 파일
    categories.csv  제품군 한 줄. block 열로 본 과제와 연습을 가른다
    details.csv     정보 국면에 한 줄씩 놓이는 특징 (제품군당 UT 3 + HE 3)
    phase2.csv      결정 국면에 놓이는 요약 문장 (제품군당 UT 1 + HE 1)
    sources.csv     stim/ 에서 그대로 복사
    brands.csv      stim/ 에서 그대로 복사

연습 제품군은 엑셀에 없다. 본 과제에 안 나오는 제품군이 하나는 있어야 해서
stim/ 에 남아 있는 바람막이를 연습 전용으로 끌어다 쓴다. 아래 PRACTICE에
그 문구가 들어 있다.
"""

from __future__ import annotations

import csv
import io
import os
import re
import shutil
import sys

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = os.path.dirname(os.path.abspath(__file__))
SRC_STIM = os.path.join(ROOT, "stim")
OUT_STIM = os.path.join(ROOT, "stim_0818")
DEFAULT_XLSX = os.path.join(
    os.path.expanduser("~"), "Downloads", "0818_product_list.xlsx"
)

# 연습 전용 제품군. 본 과제 21개에 안 들어가야 한다.
# 특징 6개는 stim/details.csv 의 windbreaker 행을 그대로 가져왔다.
# PhaseII 두 줄은 엑셀에 없어서 나머지 21개 문형에 맞춰 새로 썼다.
PRACTICE = {
    "code": "windbreaker",
    "kr": "바람막이",
    "major": "의류·잡화",
    "low": 34000,
    "high": 50000,
    "UT": ["봉제선까지 막은 방수 처리",
           "접어서 주머니에 들어가는 부피",
           "세탁기 사용 가능"],
    "HE": ["목까지 올라오는 깃",
           "도심에서 입기 좋은 실루엣",
           "다섯 가지 색상 구성"],
    "PhaseII_UT": "빈틈없는 방수 처리와 접이식 부피로 갑작스러운 비바람에도 대응 가능",
    "PhaseII_HE": "목까지 올라오는 깃과 도심에 어울리는 실루엣이 주는 단정한 인상",
}


def won(value):
    """'33,000원' 이나 33000 을 정수로."""
    digits = re.sub(r"[^0-9]", "", str(value))
    if not digits:
        raise SystemExit("가격을 읽지 못했습니다: %r" % (value,))
    return int(digits)


def cell(row, i):
    v = row[i]
    return "" if v is None else str(v).strip()


def read_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["product_list"] if "product_list" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    head = [cell(rows[0], i) for i in range(13)]
    want = ["product_EN", "product_KR", "category", "low", "high",
            "UT_1", "UT_2", "UT_3", "PhaseII_UT",
            "HE_1", "HE_2", "HE_3", "PhaseII_HE"]
    if head != want:
        raise SystemExit("열 구성이 다릅니다.\n  기대: %s\n  실제: %s" % (want, head))

    out = []
    for r in rows[1:]:
        if not cell(r, 0):
            continue
        out.append({
            "code": cell(r, 0),
            "kr": cell(r, 1),
            "major": cell(r, 2),
            "low": won(cell(r, 3)),
            "high": won(cell(r, 4)),
            "UT": [cell(r, 5), cell(r, 6), cell(r, 7)],
            "HE": [cell(r, 9), cell(r, 10), cell(r, 11)],
            "PhaseII_UT": cell(r, 8),
            "PhaseII_HE": cell(r, 12),
        })
    return out


def check(items):
    """빈칸, 코드 중복, 가격 뒤집힘만 잡는다. 문구 검수는 사람이 한다."""
    seen = set()
    for it in items:
        where = it["code"] or "(코드 없음)"
        for key in ("code", "kr", "major", "PhaseII_UT", "PhaseII_HE"):
            if not it[key]:
                raise SystemExit("%s: %s 가 비어 있습니다." % (where, key))
        for key in ("UT", "HE"):
            if any(not t for t in it[key]):
                raise SystemExit("%s: %s 문구에 빈칸이 있습니다." % (where, key))
        if it["low"] >= it["high"]:
            raise SystemExit("%s: 가격 하한이 상한보다 큽니다." % where)
        if it["code"] in seen:
            raise SystemExit("%s: 코드가 겹칩니다." % where)
        seen.add(it["code"])


def write_csv(name, header, rows):
    path = os.path.join(OUT_STIM, name)
    with io.open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    print("  %-16s %d행" % (name, len(rows)))


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        raise SystemExit("엑셀을 찾을 수 없습니다: %s" % xlsx)

    items = read_xlsx(xlsx)
    if any(it["code"] == PRACTICE["code"] for it in items):
        raise SystemExit(
            "%s 가 엑셀에 들어왔습니다. 연습 전용이라 본 과제와 겹치면 안 됩니다."
            % PRACTICE["code"]
        )
    check(items + [PRACTICE])

    os.makedirs(OUT_STIM, exist_ok=True)
    print("읽음: %s  (제품군 %d개 + 연습 1개)" % (xlsx, len(items)))

    write_csv(
        "categories.csv",
        ["category_code", "category_kr", "major_class", "price_low", "price_high",
         "block"],
        [[it["code"], it["kr"], it["major"], it["low"], it["high"], "main"]
         for it in items]
        + [[PRACTICE["code"], PRACTICE["kr"], PRACTICE["major"],
            PRACTICE["low"], PRACTICE["high"], "practice"]],
    )

    detail_rows = []
    phase2_rows = []
    for it in items + [PRACTICE]:
        for dtype in ("UT", "HE"):
            for text in it[dtype]:
                detail_rows.append([it["code"], dtype, text])
            phase2_rows.append([it["code"], dtype, it["PhaseII_" + dtype]])
    write_csv("details.csv", ["category_code", "detail_type", "text"], detail_rows)
    write_csv("phase2.csv", ["category_code", "detail_type", "text"], phase2_rows)

    for name in ("sources.csv", "brands.csv"):
        shutil.copyfile(os.path.join(SRC_STIM, name), os.path.join(OUT_STIM, name))
        print("  %-16s stim/ 에서 복사" % name)

    majors = {}
    for it in items:
        majors.setdefault(it["major"], []).append(it["kr"])
    print("\n대분류 %d개" % len(majors))
    for m, names in majors.items():
        print("  %-8s %s" % (m, ", ".join(names)))
    print("\n본 과제 %d제품군 x 반복 %d = %d시행 (정보원 6종 기준 정보원당 %.1f)"
          % (len(items), 2, len(items) * 2, len(items) * 2 / 6.0))


if __name__ == "__main__":
    main()
