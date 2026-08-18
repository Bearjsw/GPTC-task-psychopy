# -*- coding: utf-8 -*-
"""자극 CSV를 엑셀 한 파일로 묶는다.

stim/ 아래 CSV가 원본이다. 이 파일은 그것을 보기 좋게 모아 놓은 사본이라,
내용을 고칠 때는 CSV를 고치고 이걸 다시 돌리면 된다.

    python tools_make_xlsx.py
"""

from __future__ import annotations

import csv
import io
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = os.path.dirname(os.path.abspath(__file__))
STIM = os.path.join(ROOT, "stim")
OUT = os.path.join(ROOT, "GPTC_자극목록.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
UT_FILL = PatternFill("solid", fgColor="E8F0FE")
HE_FILL = PatternFill("solid", fgColor="FCE8E6")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SECTOR_KR = {
    "ELEC": "전자기기",
    "SELF-CARE": "셀프케어",
    "STA": "문구·사무",
    "CLOTHING": "의류·잡화",
}


def read(name):
    with io.open(os.path.join(STIM, name), encoding="utf-8-sig", newline="") as fh:
        return [r for r in csv.DictReader(fh)
                if any((v or "").strip() for v in r.values())]


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    cats = read("categories.csv")
    details = read("details.csv")
    sources = read("sources.csv")

    pools = {}
    for d in details:
        pools.setdefault(d["category_code"], {"UT": [], "HE": []})
        pools[d["category_code"]][d["detail_type"].upper()].append(d["text"])

    wb = Workbook()

    # ── 1. 제품군 (넓은 형식) ────────────────────────────────
    ws = wb.active
    ws.title = "제품군"
    header = ["코드", "제품군", "섹터", "가격 하한", "가격 상한",
              "실용 문구 1", "실용 문구 2", "실용 문구 3",
              "감성 문구 1", "감성 문구 2", "감성 문구 3"]
    ws.append(header)
    for c in cats:
        code = c["category_code"]
        ut = pools.get(code, {}).get("UT", []) + [""] * 3
        he = pools.get(code, {}).get("HE", []) + [""] * 3
        ws.append([code, c["category_kr"],
                   SECTOR_KR.get(c["major_class"], c["major_class"]),
                   int(c["price_low"]), int(c["price_high"])]
                  + ut[:3] + he[:3])
    style_header(ws)
    set_widths(ws, [16, 18, 12, 11, 11] + [26] * 6)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in row[3:5]:
            cell.number_format = '#,##0"원"'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        for cell in row[5:8]:
            cell.fill = UT_FILL
        for cell in row[8:11]:
            cell.fill = HE_FILL
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:K%d" % (len(cats) + 1)

    # ── 2. 문구 (긴 형식) ────────────────────────────────────
    ws = wb.create_sheet("문구")
    ws.append(["제품군", "섹터", "유형", "문구"])
    kr = {c["category_code"]: c["category_kr"] for c in cats}
    sect = {c["category_code"]: SECTOR_KR.get(c["major_class"], c["major_class"])
            for c in cats}
    for d in details:
        code = d["category_code"]
        dtype = d["detail_type"].upper()
        ws.append([kr.get(code, code), sect.get(code, ""),
                   "실용(UT)" if dtype == "UT" else "감성(HE)", d["text"]])
    style_header(ws)
    set_widths(ws, [18, 12, 12, 46])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[3].fill = UT_FILL if row[2].value.startswith("실용") else HE_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:D%d" % (len(details) + 1)

    # ── 3. 정보원 ───────────────────────────────────────────
    ws = wb.create_sheet("정보원")
    ws.append(["코드", "화면 라벨", "설문에서 부르는 이름", "소개문"])
    for s in sources:
        ws.append([s["source_code"], s["label"], s["short_name"], s["intro_text"]])
    style_header(ws)
    set_widths(ws, [14, 16, 20, 80])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row[0].row].height = 34
    ws.freeze_panes = "A2"

    # ── 4. 설계 요약 ────────────────────────────────────────
    ws = wb.create_sheet("설계 요약")
    n_cat, n_ex = len(cats), 1
    kept = n_cat - n_ex
    sets_per, reps, n_src = 3, 2, len(sources)
    trials = kept * sets_per * reps
    rows = [
        ("항목", "값", "설명"),
        ("제품군 수", n_cat, "categories.csv 행 수"),
        ("연습 전용 제품군", n_ex, "categories.csv의 block이 practice인 것"),
        ("과제에 쓰는 제품군", kept, ""),
        ("제품군당 세트", sets_per, "세트 하나가 후보 3개"),
        ("세트 반복", reps, "같은 세트를 두 번, 정보원만 바꿔서"),
        ("총 시행", trials, "%d x %d x %d" % (kept, sets_per, reps)),
        ("정보원 수", n_src, ""),
        ("정보원당 시행", trials // n_src, ""),
        ("연습 시행", 4, "연습 전용 제품군을 쓴다"),
        ("제품군당 문구", len(details) // n_cat, "실용 3 + 감성 3"),
        ("가격 범위", "%s ~ %s원" % (
            format(min(int(c["price_low"]) for c in cats), ","),
            format(max(int(c["price_high"]) for c in cats), ",")), "세트마다 뽑는다"),
        ("브랜드명", "알파벳 4~5자", "참가자마다 실행 중에 새로 뽑는다"),
    ]
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    set_widths(ws, [22, 20, 42])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[1].alignment = Alignment(horizontal="center", vertical="center")

    note = ws.cell(row=len(rows) + 3, column=1)
    note.value = ("원본은 stim/ 아래 CSV다. 내용을 고칠 때는 CSV를 고치고 "
                  "tools_make_xlsx.py를 다시 돌린다.")
    note.font = Font(italic=True, color="808080")

    wb.save(OUT)
    print("저장: %s" % OUT)
    print("  제품군 %d개, 문구 %d개, 정보원 %d종" % (len(cats), len(details), len(sources)))


if __name__ == "__main__":
    main()
